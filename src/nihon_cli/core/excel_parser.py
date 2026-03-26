"""Parser for Excel vocabulary files (e.g. まるごと textbook exports).

Handles varying column layouts by detecting headers dynamically.
"""

from pathlib import Path
from typing import List, Optional

from nihon_cli.core.vocabulary import VocabularyItem


# Map Japanese word type labels to internal vocab_type values
_VOCAB_TYPE_MAP = {
    "名詞": "noun",
    "動詞": "noun",
    "い形容詞": "adjective",
    "な形容詞": "adjective",
    "副詞": "noun",
    "助詞": "noun",
    "接頭語": "noun",
    "接尾語": "noun",
    "感動詞": "noun",
    "接続詞": "noun",
    "連体詞": "noun",
    "あいさつ": "pattern",
}


def _detect_vocab_type(raw_type: Optional[str]) -> Optional[str]:
    """Map a Japanese word type label like '01 名詞' to an internal type."""
    if not raw_type or not isinstance(raw_type, str):
        return None
    for key, value in _VOCAB_TYPE_MAP.items():
        if key in raw_type:
            return value
    return None


class ExcelVocabParser:
    """Parse vocabulary from まるごと-style Excel workbooks."""

    def parse(self, file_path: Path) -> dict[str, List[VocabularyItem]]:
        """Parse all themed sheets from an Excel file.

        Args:
            file_path: Path to the .xlsx file

        Returns:
            Dict mapping sheet name to list of VocabularyItem objects.
            The first sheet (全単語) is skipped since themed sheets
            contain the same data organized by topic.

        Raises:
            FileNotFoundError: If the file does not exist
            ValueError: If no vocabulary could be parsed
        """
        import openpyxl

        if not file_path.exists():
            raise FileNotFoundError(f"Datei nicht gefunden: {file_path}")

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        result: dict[str, List[VocabularyItem]] = {}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            items = self._parse_sheet(ws)
            if items:
                result[sheet_name] = items

        wb.close()

        if not result:
            raise ValueError("Keine Vokabeln in der Excel-Datei gefunden.")

        return result

    def _parse_sheet(self, ws) -> List[VocabularyItem]:
        """Parse a single worksheet."""
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            return []

        # Find header row by looking for '語彙' or 'ドイツ語訳'
        header_idx = None
        for i, row in enumerate(rows[:5]):
            row_strs = [str(c) if c else "" for c in row]
            if any("語彙" in s for s in row_strs) or any("ドイツ語訳" in s for s in row_strs):
                header_idx = i
                break

        if header_idx is None:
            return []

        headers = [str(c).strip() if c else "" for c in rows[header_idx]]

        # Detect column indices
        col_kana = self._find_col(headers, "かな")
        col_kanji = self._find_col(headers, "漢字")
        col_german = self._find_col(headers, "ドイツ語訳")
        col_type = self._find_col(headers, "品詞")

        if col_kana is None or col_german is None:
            return []

        items: List[VocabularyItem] = []
        for row in rows[header_idx + 1:]:
            if not row or len(row) <= max(col_kana, col_german):
                continue

            kana = str(row[col_kana]).strip() if row[col_kana] else None
            if not kana or kana == "None":
                continue

            kanji = None
            if col_kanji is not None and len(row) > col_kanji and row[col_kanji]:
                kanji_raw = str(row[col_kanji]).strip()
                if kanji_raw and kanji_raw != kana and kanji_raw != "None":
                    kanji = kanji_raw

            german = str(row[col_german]).strip() if row[col_german] else None
            if not german or german == "None":
                continue

            # Build japanese vocab list: kanji form + kana reading
            japanese = [kanji] if kanji else [kana]
            if kanji and kanji != kana:
                japanese.append(kana)

            # German may have multiple meanings separated by comma
            german_list = [g.strip() for g in german.split(",") if g.strip()]

            raw_type = None
            if col_type is not None and len(row) > col_type:
                raw_type = str(row[col_type]) if row[col_type] else None

            vocab_type = _detect_vocab_type(raw_type)

            items.append(VocabularyItem(
                id=0,
                japanese_vocab=japanese,
                german_vocab=german_list,
                source_file=None,
                upload_tag=None,
                vocab_type=vocab_type,
            ))

        return items

    @staticmethod
    def _find_col(headers: List[str], keyword: str) -> Optional[int]:
        """Find the column index containing a keyword."""
        for i, h in enumerate(headers):
            if keyword in h:
                return i
        return None
